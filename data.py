
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("WOS_data.xlsx")
current_ws = wb["source"]
record_length = 6032
title_column = get_column_letter(1)
year_column = get_column_letter(2)
wb.create_sheet('ion-exchange')
new_ws = wb['ion-exchange']
words = ["exchange resin", "ion-exchange", "ion exchange", "resin"]
# words = ["zero-valent", "zero valent"]
# words = ["adsorption"]
# words = ["electrochemical", "electrochemical reduction", "electrode", "electrodes"]
# words = ["catalytic", "catalyst", "catalysis", "biocatalyst", "biocatalytic", "biocatalysis", "electrocatalyst", 
#          "electrocatalysis", "electrocatalytic", "photocatalyst", "photocatalytic", "photocatalysis", "electrochemical", 
#          "electroreduction", "electrode", "electrodes"]
keywords = set(words)


def my_function(title_column, year_column, keywords):
    for row in range(1, record_length+1):
     title1 = current_ws[title_column + str(row)].value
     title = title1.lower()
     titleParts = set(title.split())
     if keywords.intersection(titleParts):
        year = current_ws[year_column + str(row)].value
        new_ws.append([title,year])
    wb.save('WOS_data.xlsx')


my_function(title_column, year_column, keywords)

